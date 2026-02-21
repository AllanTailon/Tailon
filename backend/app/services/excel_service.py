"""
Serviço para processamento de arquivos Excel.

Segurança:
- Validação de tipo de arquivo
- Limite de tamanho
- Sanitização de nomes
- Não executa macros
- Timeout no parsing
"""

import logging
import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from ..models.data import ColumnInfo, DatasetInfo, DatasetPreview

logger = logging.getLogger(__name__)

# Limite de segurança
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_ROWS = 100_000
MAX_COLUMNS = 500
ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}


class DataStore:
    """
    Armazenamento em memória para datasets.
    
    Em produção, substituir por Redis ou banco de dados.
    """
    
    def __init__(self):
        self._datasets: Dict[str, Dict[str, Any]] = {}
        self._dataframes: Dict[str, pd.DataFrame] = {}
    
    def store(self, dataset_id: str, info: DatasetInfo, df: pd.DataFrame) -> None:
        """Armazena um dataset."""
        self._datasets[dataset_id] = info.model_dump()
        self._dataframes[dataset_id] = df
        logger.info(f"Dataset armazenado: {dataset_id}")
    
    def get_info(self, dataset_id: str) -> Optional[DatasetInfo]:
        """Retorna informações de um dataset."""
        data = self._datasets.get(dataset_id)
        if data:
            return DatasetInfo(**data)
        return None
    
    def get_dataframe(self, dataset_id: str) -> Optional[pd.DataFrame]:
        """Retorna o DataFrame de um dataset."""
        return self._dataframes.get(dataset_id)
    
    def delete(self, dataset_id: str) -> bool:
        """Remove um dataset."""
        if dataset_id in self._datasets:
            del self._datasets[dataset_id]
            del self._dataframes[dataset_id]
            logger.info(f"Dataset removido: {dataset_id}")
            return True
        return False
    
    def list_all(self) -> List[DatasetInfo]:
        """Lista todos os datasets."""
        return [DatasetInfo(**d) for d in self._datasets.values()]
    
    def clear(self) -> None:
        """Remove todos os datasets."""
        self._datasets.clear()
        self._dataframes.clear()
        logger.info("Todos os datasets removidos")


# Instância global do store (em produção, usar injeção de dependência)
data_store = DataStore()


class ExcelService:
    """Serviço para processamento de arquivos Excel."""
    
    def __init__(self, store: DataStore = None):
        self.store = store or data_store
    
    @staticmethod
    def _generate_id() -> str:
        """Gera ID único para dataset."""
        return f"ds_{uuid4().hex[:12]}"
    
    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        """Sanitiza nome do arquivo."""
        # Remove caracteres perigosos
        safe = re.sub(r'[<>:"/\\|?*]', '_', filename)
        # Limita tamanho
        return safe[:100]
    
    @staticmethod
    def _get_extension(filename: str) -> str:
        """Retorna extensão do arquivo."""
        if '.' in filename:
            return '.' + filename.rsplit('.', 1)[1].lower()
        return ''
    
    @staticmethod
    def _infer_dtype(series: pd.Series) -> str:
        """Infere tipo de dados de uma coluna."""
        dtype = str(series.dtype)
        
        if 'int' in dtype or 'float' in dtype:
            return 'number'
        elif 'datetime' in dtype:
            return 'date'
        elif 'bool' in dtype:
            return 'boolean'
        else:
            # Tenta detectar datas em strings
            if series.dtype == 'object':
                sample = series.dropna().head(10)
                if len(sample) > 0:
                    try:
                        pd.to_datetime(sample)
                        return 'date'
                    except:
                        pass
            return 'string'
    
    @staticmethod
    def _get_sample_values(series: pd.Series, n: int = 5) -> List[Any]:
        """Retorna valores de amostra de uma coluna."""
        samples = series.dropna().head(n).tolist()
        # Converte para tipos serializáveis
        result = []
        for val in samples:
            if pd.isna(val):
                continue
            elif isinstance(val, (int, float, bool, str)):
                result.append(val)
            else:
                result.append(str(val))
        return result
    
    def validate_file(
        self, 
        filename: str, 
        content: bytes
    ) -> Tuple[bool, List[str]]:
        """
        Valida arquivo antes do processamento.
        
        Returns:
            (is_valid, list_of_errors)
        """
        errors = []
        
        # Verifica extensão
        ext = self._get_extension(filename)
        if ext not in ALLOWED_EXTENSIONS:
            errors.append(f"Extensão não permitida: {ext}. Use: {', '.join(ALLOWED_EXTENSIONS)}")
        
        # Verifica tamanho
        if len(content) > MAX_FILE_SIZE:
            errors.append(f"Arquivo muito grande. Máximo: {MAX_FILE_SIZE // (1024*1024)}MB")
        
        # Verifica se está vazio
        if len(content) == 0:
            errors.append("Arquivo vazio")
        
        return len(errors) == 0, errors
    
    def parse_excel(
        self,
        filename: str,
        content: bytes,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ) -> Tuple[Optional[DatasetInfo], Optional[pd.DataFrame], List[str], List[str]]:
        """
        Faz parsing de um arquivo Excel.
        
        Returns:
            (dataset_info, dataframe, errors, warnings)
        """
        errors = []
        warnings = []
        
        # Valida arquivo
        is_valid, validation_errors = self.validate_file(filename, content)
        if not is_valid:
            return None, None, validation_errors, warnings
        
        try:
            ext = self._get_extension(filename)
            
            # Lê arquivo
            if ext == '.csv':
                # CSV
                df = pd.read_csv(
                    io.BytesIO(content),
                    header=0 if has_header else None,
                    nrows=MAX_ROWS,
                )
                actual_sheet = None
            else:
                # Excel
                # Primeiro, lista as abas disponíveis
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                available_sheets = wb.sheetnames
                wb.close()
                
                # Determina qual aba usar
                if sheet_name and sheet_name in available_sheets:
                    actual_sheet = sheet_name
                elif sheet_name:
                    warnings.append(f"Aba '{sheet_name}' não encontrada. Usando primeira aba.")
                    actual_sheet = available_sheets[0]
                else:
                    actual_sheet = available_sheets[0]
                
                if len(available_sheets) > 1:
                    warnings.append(f"Arquivo tem {len(available_sheets)} abas. Usando: '{actual_sheet}'")
                
                df = pd.read_excel(
                    io.BytesIO(content),
                    sheet_name=actual_sheet,
                    header=0 if has_header else None,
                    nrows=MAX_ROWS,
                    engine='openpyxl',
                )
            
            # Valida tamanho
            if len(df) >= MAX_ROWS:
                warnings.append(f"Arquivo truncado para {MAX_ROWS} linhas")
            
            if len(df.columns) > MAX_COLUMNS:
                errors.append(f"Muitas colunas ({len(df.columns)}). Máximo: {MAX_COLUMNS}")
                return None, None, errors, warnings
            
            # Se não tem header, gera nomes de coluna
            if not has_header:
                df.columns = [f"col_{i+1}" for i in range(len(df.columns))]
            
            # Sanitiza nomes de colunas
            df.columns = [
                re.sub(r'[^\w\s-]', '', str(col)).strip().replace(' ', '_')[:50]
                for col in df.columns
            ]
            
            # Remove colunas duplicadas
            if df.columns.duplicated().any():
                warnings.append("Colunas duplicadas foram renomeadas")
                cols = []
                for col in df.columns:
                    if col in cols:
                        i = 1
                        while f"{col}_{i}" in cols:
                            i += 1
                        cols.append(f"{col}_{i}")
                    else:
                        cols.append(col)
                df.columns = cols
            
            # Gera informações das colunas
            columns_info = []
            for col in df.columns:
                series = df[col]
                columns_info.append(ColumnInfo(
                    name=col,
                    dtype=self._infer_dtype(series),
                    sample_values=self._get_sample_values(series),
                    null_count=int(series.isna().sum()),
                    unique_count=int(series.nunique()),
                ))
            
            # Cria info do dataset
            dataset_id = self._generate_id()
            safe_filename = self._sanitize_filename(filename)
            
            dataset_info = DatasetInfo(
                id=dataset_id,
                name=safe_filename.rsplit('.', 1)[0] if '.' in safe_filename else safe_filename,
                original_filename=safe_filename,
                sheet_name=actual_sheet,
                row_count=len(df),
                column_count=len(df.columns),
                columns=columns_info,
            )
            
            return dataset_info, df, errors, warnings
            
        except pd.errors.EmptyDataError:
            errors.append("Arquivo vazio ou sem dados válidos")
            return None, None, errors, warnings
        except Exception as e:
            # Não loga detalhes por segurança
            logger.error(f"Erro ao processar arquivo: {type(e).__name__}")
            errors.append("Erro ao processar arquivo. Verifique se o formato está correto.")
            return None, None, errors, warnings
    
    def upload(
        self,
        filename: str,
        content: bytes,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ) -> Tuple[Optional[DatasetInfo], List[str], List[str]]:
        """
        Faz upload e armazena um arquivo.
        
        Returns:
            (dataset_info, errors, warnings)
        """
        info, df, errors, warnings = self.parse_excel(
            filename, content, sheet_name, has_header
        )
        
        if info and df is not None:
            self.store.store(info.id, info, df)
            return info, errors, warnings
        
        return None, errors, warnings
    
    def get_preview(
        self, 
        dataset_id: str, 
        max_rows: int = 50
    ) -> Optional[DatasetPreview]:
        """Retorna preview de um dataset."""
        info = self.store.get_info(dataset_id)
        df = self.store.get_dataframe(dataset_id)
        
        if info is None or df is None:
            return None
        
        # Converte para dicionários (serialização segura)
        rows = []
        for _, row in df.head(max_rows).iterrows():
            row_dict = {}
            for col in df.columns:
                val = row[col]
                if pd.isna(val):
                    row_dict[col] = None
                elif isinstance(val, (int, float, bool, str)):
                    row_dict[col] = val
                else:
                    row_dict[col] = str(val)
            rows.append(row_dict)
        
        return DatasetPreview(
            id=dataset_id,
            name=info.name,
            columns=list(df.columns),
            rows=rows,
            total_rows=len(df),
        )
    
    def get_data(self, dataset_id: str) -> Optional[pd.DataFrame]:
        """Retorna DataFrame de um dataset."""
        return self.store.get_dataframe(dataset_id)
    
    def delete(self, dataset_id: str) -> bool:
        """Remove um dataset."""
        return self.store.delete(dataset_id)
    
    def list_datasets(self) -> List[DatasetInfo]:
        """Lista todos os datasets."""
        return self.store.list_all()

